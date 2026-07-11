from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill('solid', fgColor='DECDA6')


def _display_user(user):
    return user.get_full_name() or user.username if user else ''


def _display_client(client):
    return str(client) if client else ''


def _date(value):
    if not value:
        return ''
    return value.strftime('%Y-%m-%d') if hasattr(value, 'strftime') else value


def _datetime(value):
    if not value:
        return ''
    return value.strftime('%Y-%m-%d %H:%M') if hasattr(value, 'strftime') else value


def _money(value):
    return float(value or 0)


def _write_sheet(title, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    _fill_sheet(sheet, headers, rows)
    return _to_bytes(workbook)


def _fill_sheet(sheet, headers, rows):
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for row in rows:
        sheet.append(row)
    for index, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for cell in sheet[get_column_letter(index)]:
            max_length = max(max_length, len(str(cell.value or '')))
        sheet.column_dimensions[get_column_letter(index)].width = min(max_length + 3, 45)
    sheet.freeze_panes = 'A2'


def _to_bytes(workbook):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_clients(queryset):
    headers = ['ID', 'ФИО ученика', 'Родитель', 'Телефон', 'Менеджер', 'Статус', 'Дата создания', 'Комментарий']
    rows = [
        [
            client.id,
            str(client),
            client.parent_name,
            client.phone,
            _display_user(client.manager),
            'Активен' if client.is_active else 'Неактивен',
            _datetime(client.created_at),
            client.notes,
        ]
        for client in queryset
    ]
    return _write_sheet('Клиенты', headers, rows)


def export_subscriptions(queryset):
    headers = ['ID', 'Клиент', 'Телефон', 'Название', 'Всего занятий', 'Осталось', 'Использовано', 'Прогресс %', 'Цена', 'Оплачено', 'Статус', 'Дата начала', 'Дата окончания', 'Продление']
    rows = []
    for item in queryset:
        used = max((item.total_visits or 0) - (item.remaining_visits or 0), 0)
        progress = round(used / item.total_visits * 100, 2) if item.total_visits else 0
        rows.append([
            item.id,
            _display_client(item.client),
            item.client.phone if item.client else '',
            item.title,
            item.total_visits,
            item.remaining_visits,
            used,
            progress,
            _money(item.price),
            _money(item.paid_amount),
            item.status,
            _date(item.start_date),
            _date(item.end_date),
            '',
        ])
    return _write_sheet('Абонементы', headers, rows)


def export_visits(queryset):
    headers = ['ID', 'Дата', 'Клиент', 'Телефон', 'Абонемент', 'Учитель', 'Урок', 'Статус', 'Комментарий', 'Списано занятие']
    rows = [
        [
            visit.id,
            _datetime(visit.visited_at),
            _display_client(visit.client),
            visit.client.phone if visit.client else '',
            visit.subscription.title if visit.subscription else '',
            _display_user(visit.teacher),
            str(visit.lesson) if visit.lesson else '',
            visit.status,
            visit.notes,
            'Да' if visit.lesson_deducted else 'Нет',
        ]
        for visit in queryset
    ]
    return _write_sheet('Посещения', headers, rows)


def export_finance(queryset):
    headers = ['ID', 'Дата', 'Тип', 'Источник', 'Клиент', 'Сумма', 'Метод оплаты', 'Создал', 'Описание']
    rows = [
        [
            item.id,
            _datetime(item.paid_at),
            item.transaction_type,
            item.source,
            _display_client(item.client),
            _money(item.amount),
            item.payment_method_name or (item.payment_method.name if item.payment_method else ''),
            _display_user(item.created_by),
            item.comment,
        ]
        for item in queryset
    ]
    return _write_sheet('Финансы', headers, rows)


def export_trials(queryset):
    headers = ['ID', 'Ученик', 'Родитель', 'Телефон', 'Клиент', 'Менеджер', 'Этап', 'Дата пробного', 'Дата оплаты', 'Сумма', 'Метод оплаты', 'Комментарий']
    rows = [
        [
            item.id,
            _display_client(item.client),
            item.client.parent_name if item.client else '',
            item.client.phone if item.client else '',
            _display_client(item.client),
            _display_user(item.manager),
            item.status,
            _datetime(item.scheduled_at),
            _date(item.payment_date),
            _money(item.price),
            '',
            item.notes,
        ]
        for item in queryset
    ]
    return _write_sheet('Пробники', headers, rows)


def export_master_classes(queryset):
    headers = ['ID', 'Ученик', 'Телефон', 'Клиент', 'Менеджер', 'Куратор', 'Предмет', 'Этап', 'Статус', 'Дата МК', 'Дата оплаты', 'Сумма оплаты', 'Метод оплаты', 'Комментарий']
    rows = []
    for item in queryset:
        client = item.participants.first()
        rows.append([
            item.id,
            _display_client(client),
            client.phone if client else '',
            _display_client(client),
            _display_user(item.manager),
            _display_user(item.teacher),
            item.title,
            item.stage,
            item.stage,
            _datetime(item.starts_at),
            _date(item.payment_date),
            _money(item.payment_amount),
            '',
            item.description,
        ])
    return _write_sheet('МК', headers, rows)


def export_groups(queryset):
    headers = ['ID', 'Название', 'Предмет', 'Учитель', 'Менеджер', 'Статус', 'Кол-во учеников', 'Дата начала', 'Дата окончания', 'Описание']
    rows = [
        [
            item.id,
            item.name,
            item.subject.name if item.subject else '',
            _display_user(item.teacher),
            _display_user(item.manager),
            item.status,
            item.memberships.filter(status='active').count(),
            _date(item.start_date),
            _date(item.end_date),
            item.description,
        ]
        for item in queryset
    ]
    return _write_sheet('Группы', headers, rows)


def export_lessons(queryset):
    headers = ['ID', 'Дата', 'Время', 'Группа', 'Предмет', 'Учитель', 'Кабинет', 'Статус', 'Тема', 'Посещений всего', 'Пришли', 'Пропустили', 'Комментарий']
    rows = [
        [
            item.id,
            _date(item.lesson_date),
            f'{item.start_time:%H:%M} - {item.end_time:%H:%M}',
            item.group.name if item.group else '',
            item.subject.name if item.subject else '',
            _display_user(item.teacher),
            item.room.name if item.room else '',
            item.status,
            item.topic,
            item.visits.count(),
            item.visits.filter(status='attended').count(),
            item.visits.filter(status='missed').count(),
            item.comment,
        ]
        for item in queryset
    ]
    return _write_sheet('Уроки', headers, rows)


def export_summary_report(data):
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = [
        ('Финансы по дням', ['Дата', 'Доход', 'Расход', 'Баланс'], data.get('daily_finance', []), ['date', 'income', 'expense', 'balance']),
        ('Доходы по источникам', ['Источник', 'Кол-во', 'Сумма'], data.get('income_by_source', []), ['source_display', 'count', 'amount']),
        ('Продажи менеджеров', ['Менеджер', 'Пробники', 'Купили', 'Конверсия %', 'МК', 'Купили МК', 'Конверсия МК %', 'Доход'], data.get('sales_by_manager', []), ['manager_name', 'trials_total', 'trials_bought', 'trials_conversion', 'mk_total', 'mk_bought', 'mk_conversion', 'income']),
        ('Посещаемость групп', ['Группа', 'Уроков', 'Учеников', 'Пришли', 'Пропустили', 'Посещаемость %'], data.get('attendance_by_group', []), ['group_name', 'lessons_count', 'students_count', 'attended', 'missed', 'attendance_rate']),
        ('Посещаемость учителей', ['Учитель', 'Уроков', 'Пришли', 'Пропустили', 'Посещаемость %'], data.get('attendance_by_teacher', []), ['teacher_name', 'lessons_count', 'attended', 'missed', 'attendance_rate']),
        ('Заканчиваются', ['Клиент', 'Телефон', 'Абонемент', 'Остаток', 'Всего', 'Дата окончания', 'Статус'], data.get('ending_subscriptions', []), ['client_name', 'client_phone', 'title', 'lessons_left', 'lessons_total', 'end_date', 'status']),
        ('Низкая посещаемость', ['Клиент', 'Телефон', 'Группа', 'Пришли', 'Пропустили', 'Посещаемость %'], data.get('low_attendance_clients', []), ['client_name', 'client_phone', 'group_name', 'attended', 'missed', 'attendance_rate']),
    ]
    for title, headers, rows, keys in sheets:
        sheet = workbook.create_sheet(title[:31])
        _fill_sheet(sheet, headers, [[row.get(key, '') for key in keys] for row in rows])
    return _to_bytes(workbook)
