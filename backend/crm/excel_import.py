from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation

from django.contrib.auth import get_user_model
from django.utils import timezone
from openpyxl import load_workbook

from .models import Client, FinanceTransaction, MasterClass, PaymentMethod, Subscription, Trial, Visit


SHEET_CLIENTS = 'Клиенты'
SHEET_SUBSCRIPTIONS = 'Абонемент'
SHEET_TRIALS = 'Пробники'
SHEET_MASTER_CLASSES = 'МК'
SHEET_VISITS = 'Посещения'

EMPTY_RESULT = {
    'created': {
        'clients': 0,
        'subscriptions': 0,
        'trials': 0,
        'master_classes': 0,
        'visits': 0,
    },
    'skipped': 0,
    'warnings': [],
}


def import_excel(uploaded_file, user=None):
    result = {
        'created': EMPTY_RESULT['created'].copy(),
        'skipped': 0,
        'warnings': [],
    }

    workbook = load_workbook(uploaded_file, data_only=True)
    known_sheets = {
        SHEET_CLIENTS: _import_clients,
        SHEET_SUBSCRIPTIONS: _import_subscriptions,
        SHEET_TRIALS: _import_trials,
        SHEET_MASTER_CLASSES: _import_master_classes,
        SHEET_VISITS: _import_visits,
    }

    imported_any = False
    for sheet_name, importer in known_sheets.items():
        if sheet_name not in workbook.sheetnames:
            continue
        imported_any = True
        importer(workbook[sheet_name], result, user)

    if not imported_any:
        result['warnings'].append(
            'Не найдено ни одного поддерживаемого листа: Клиенты, Абонемент, Пробники, МК, Посещения.'
        )

    return result


def _import_clients(sheet, result, user):
    for row_number, row in _iter_dict_rows(sheet):
        try:
            full_name = _get(row, 'фио', 'клиент', 'имя')
            phone = _get(row, 'конт', 'контакты', 'телефон')
            if not full_name and not phone:
                _skip(result, sheet.title, row_number, 'нет ФИО или телефона')
                continue
            _, created = _get_or_create_client(
                full_name=full_name,
                phone=phone,
                parent_name=_get(row, 'родитель'),
                manager=_find_user(_get(row, 'менеджер')),
                notes=_get(row, 'комментарий', 'заметки'),
            )
            if created:
                result['created']['clients'] += 1
        except Exception as exc:
            _warn(result, sheet.title, row_number, str(exc))


def _import_subscriptions(sheet, result, user):
    for row_number, row in _iter_dict_rows(sheet):
        try:
            full_name = _get(row, 'фио')
            phone = _get(row, 'конт', 'контакты', 'телефон')
            if not full_name and not phone:
                _skip(result, sheet.title, row_number, 'нет ФИО или телефона')
                continue

            client, client_created = _get_or_create_client(full_name=full_name, phone=phone)
            if client_created:
                result['created']['clients'] += 1

            title = _text(_get(row, 'вид абон', 'вид абонемента')) or 'Абонемент'
            lessons_left = _int(_get(row, 'остаток занятий', 'lessons_left'), default=0)
            lessons_total = _subscription_total(title, lessons_left)
            start_date = _parse_date(_get(row, 'дата начала')) or _parse_date(_get(row, 'дата покупки'))
            if not start_date:
                _skip(result, sheet.title, row_number, 'нет даты начала или даты покупки')
                continue

            Subscription.objects.create(
                client=client,
                title=title,
                purchase_date=_parse_date(_get(row, 'дата покупки')),
                start_date=start_date,
                end_date=_parse_date(_get(row, 'дата окончания', 'продление')),
                total_visits=lessons_total,
                remaining_visits=lessons_left,
                status=_subscription_status(_get(row, 'статус')),
            )
            result['created']['subscriptions'] += 1
        except Exception as exc:
            _warn(result, sheet.title, row_number, str(exc))


def _import_trials(sheet, result, user):
    for row_number, row in _iter_dict_rows(sheet):
        try:
            full_name = _get(row, 'фио')
            phone = _get(row, 'контакты', 'конт', 'телефон')
            scheduled_at = _parse_datetime(_get(row, 'дата пробного', 'дата'))
            if not scheduled_at:
                _skip(result, sheet.title, row_number, 'нет даты пробного')
                continue

            client, client_created = _get_or_create_client(
                full_name=full_name,
                phone=phone,
                parent_name=_get(row, 'родитель'),
                manager=_find_user(_get(row, 'менеджер')),
                notes=_get(row, 'комментарий'),
            )
            if client_created:
                result['created']['clients'] += 1

            Trial.objects.create(
                client=client,
                manager=_find_user(_get(row, 'менеджер')),
                scheduled_at=scheduled_at,
                status=_trial_stage(_get(row, 'этап/статус', 'статус', 'этап')),
                bought_subscription=_trial_stage(_get(row, 'этап/статус', 'статус', 'этап')) == Trial.Status.BOUGHT,
                notes=_join_notes(_get(row, 'напоминание'), _get(row, 'комментарий')),
            )
            result['created']['trials'] += 1
        except Exception as exc:
            _warn(result, sheet.title, row_number, str(exc))


def _import_master_classes(sheet, result, user):
    for row_number, row in _iter_dict_rows(sheet):
        try:
            full_name = _get(row, 'фио')
            phone = _get(row, 'телефон', 'конт', 'контакты')
            starts_at = _parse_datetime(_get(row, 'дата мк', 'дата'))
            if not starts_at:
                _skip(result, sheet.title, row_number, 'нет даты МК')
                continue

            client, client_created = _get_or_create_client(full_name=full_name, phone=phone)
            if client_created:
                result['created']['clients'] += 1

            payment_amount = _decimal(_get(row, 'сумма оплаты'), default=0)
            payment_date = _parse_date(_get(row, 'дата оплаты'))
            master_class = MasterClass.objects.create(
                title=_text(_get(row, 'предмет')) or 'Мастер-класс',
                description=_join_notes(_get(row, 'whatsapp напоминание'), _get(row, 'комментарий')),
                manager=_find_user(_get(row, 'менеджер')),
                teacher=_find_user(_get(row, 'куратор')),
                starts_at=starts_at,
                stage=_master_class_stage(_get(row, 'этап', 'статус')),
                payment_amount=payment_amount,
                payment_date=payment_date,
                price=payment_amount,
                capacity=1,
            )
            master_class.participants.add(client)
            if payment_amount > 0:
                method_text = _text(_get(row, 'метод оплаты'))
                payment_method = PaymentMethod.objects.filter(name__iexact=method_text).first() if method_text else None
                finance_transaction = FinanceTransaction.objects.create(
                    transaction_type=FinanceTransaction.Type.INCOME,
                    amount=payment_amount,
                    source='master_class',
                    payment_method=payment_method,
                    payment_method_name=payment_method.name if payment_method else method_text,
                    client=client,
                    created_by=user if getattr(user, 'is_authenticated', False) else None,
                    paid_at=_datetime_from_date(payment_date),
                    comment='Оплата МК',
                )
                master_class.finance_transaction = finance_transaction
                master_class.save(update_fields=('finance_transaction', 'updated_at'))
            result['created']['master_classes'] += 1
        except Exception as exc:
            _warn(result, sheet.title, row_number, str(exc))


def _import_visits(sheet, result, user):
    for row_number, row in _iter_dict_rows(sheet):
        try:
            full_name = _find_value(row, 'фио', 'клиент', 'ученик')
            visited_at = _parse_datetime(_find_value(row, 'дата занятия', 'дата', 'занятие', 'visited_at'))
            if not full_name or not visited_at:
                _skip(result, sheet.title, row_number, 'недостаточно данных для посещения')
                continue

            client, client_created = _get_or_create_client(full_name=full_name)
            if client_created:
                result['created']['clients'] += 1

            subscription = client.subscriptions.order_by('-start_date', '-created_at').first()
            Visit.objects.create(
                client=client,
                subscription=subscription,
                visited_at=visited_at,
                status=_visit_status(_find_value(row, 'статус', 'посещение')),
                notes=_visit_notes(row),
            )
            lessons_left = _int(_find_value(row, 'остаток занятий', 'остаток'), default=None)
            if subscription and lessons_left is not None:
                subscription.remaining_visits = max(lessons_left, 0)
                subscription.save(update_fields=('remaining_visits', 'updated_at'))
            result['created']['visits'] += 1
        except Exception as exc:
            _warn(result, sheet.title, row_number, str(exc))


def _iter_dict_rows(sheet):
    header_row = _find_header_row(sheet)
    if not header_row:
        return
    headers = [_normalize_header(cell.value) for cell in sheet[header_row]]
    for row_number in range(header_row + 1, sheet.max_row + 1):
        values = [cell.value for cell in sheet[row_number]]
        if all(_is_blank(value) for value in values):
            continue
        yield row_number, {headers[index]: values[index] for index in range(min(len(headers), len(values))) if headers[index]}


def _find_header_row(sheet):
    for row_number in range(1, min(sheet.max_row, 10) + 1):
        values = [_normalize_header(cell.value) for cell in sheet[row_number]]
        if any(value in values for value in ('фио', 'конт', 'контакты', 'телефон', 'дата мк', 'дата пробного')):
            return row_number
    return 1 if sheet.max_row else None


def _normalize_header(value):
    return ' '.join(_text(value).replace('\n', ' ').lower().split())


def _get(row, *keys):
    for key in keys:
        value = row.get(_normalize_header(key))
        if not _is_blank(value):
            return value
    return None


def _find_value(row, *needles):
    for key, value in row.items():
        if _is_blank(value):
            continue
        normalized_key = _normalize_header(key)
        if any(_normalize_header(needle) in normalized_key for needle in needles):
            return value
    return None


def _get_or_create_client(*, full_name=None, phone=None, parent_name=None, manager=None, notes=None):
    full_name = _text(full_name)
    phone = _text(phone)
    first_name, last_name = _split_name(full_name)

    client = None
    if phone:
        client = Client.objects.filter(phone=phone).first()
    if not client and full_name:
        client = Client.objects.filter(first_name=first_name, last_name=last_name).first()
    if client:
        update_fields = []
        if phone and not client.phone:
            client.phone = phone
            update_fields.append('phone')
        if parent_name and not client.parent_name:
            client.parent_name = _text(parent_name)
            update_fields.append('parent_name')
        if manager and not client.manager_id:
            client.manager = manager
            update_fields.append('manager')
        if notes and not client.notes:
            client.notes = _text(notes)
            update_fields.append('notes')
        if update_fields:
            update_fields.append('updated_at')
            client.save(update_fields=update_fields)
        return client, False

    client = Client.objects.create(
        first_name=first_name or phone or 'Без имени',
        last_name=last_name,
        parent_name=_text(parent_name),
        phone=phone,
        manager=manager,
        notes=_text(notes),
    )
    return client, True


def _split_name(full_name):
    parts = _text(full_name).split()
    if not parts:
        return '', ''
    return parts[0], ' '.join(parts[1:])


def _find_user(value):
    text = _text(value)
    if not text:
        return None
    User = get_user_model()
    user = User.objects.filter(username__iexact=text).first()
    if user:
        return user
    parts = text.split()
    if len(parts) >= 2:
        user = User.objects.filter(first_name__iexact=parts[0], last_name__iexact=' '.join(parts[1:])).first()
        if user:
            return user
    return User.objects.filter(first_name__iexact=text).first() or User.objects.filter(last_name__iexact=text).first()


def _subscription_total(title, lessons_left):
    normalized = _text(title).upper().replace(' ', '')
    if 'AB-4' in normalized or 'AB4' in normalized:
        return 4
    if 'AB-8' in normalized or 'AB8' in normalized:
        return 8
    return max(lessons_left or 0, 0)


def _subscription_status(value):
    text = _normalize_status(value)
    if text in ('пауза', 'pause', 'paused', 'заморожен', 'заморозка'):
        return Subscription.Status.PAUSED
    if text in ('истек', 'истёк', 'expired', 'закончился'):
        return Subscription.Status.EXPIRED
    if text in ('отменен', 'отменён', 'cancelled', 'canceled'):
        return Subscription.Status.CANCELLED
    return Subscription.Status.ACTIVE


def _trial_stage(value):
    text = _normalize_status(value)
    mapping = {
        'лид': Trial.Status.LEAD,
        'lead': Trial.Status.LEAD,
        'записался на пробный': Trial.Status.BOOKED,
        'booked': Trial.Status.BOOKED,
        'прошел пробный': Trial.Status.ATTENDED,
        'прошёл пробный': Trial.Status.ATTENDED,
        'attended': Trial.Status.ATTENDED,
        'купил абонемент': Trial.Status.BOUGHT,
        'bought': Trial.Status.BOUGHT,
        'не купил': Trial.Status.LOST,
        'lost': Trial.Status.LOST,
        'новый': Trial.Status.NEW,
        'new': Trial.Status.NEW,
        'запланирован': Trial.Status.SCHEDULED,
        'scheduled': Trial.Status.SCHEDULED,
        'завершен': Trial.Status.COMPLETED,
        'завершён': Trial.Status.COMPLETED,
        'completed': Trial.Status.COMPLETED,
        'отменен': Trial.Status.CANCELLED,
        'отменён': Trial.Status.CANCELLED,
        'cancelled': Trial.Status.CANCELLED,
        'canceled': Trial.Status.CANCELLED,
    }
    return mapping.get(text, Trial.Status.NEW)


def _master_class_stage(value):
    text = _normalize_status(value)
    if text in ('завершен', 'завершён', 'прошел', 'прошёл', 'completed', 'attended'):
        return MasterClass.Stage.COMPLETED
    if text in ('отменен', 'отменён', 'cancelled', 'canceled', 'lost'):
        return MasterClass.Stage.CANCELLED
    return MasterClass.Stage.PLANNED


def _visit_status(value):
    text = _normalize_status(value)
    if text in ('посетил', 'был', 'attended', 'прошел', 'прошёл'):
        return Visit.Status.ATTENDED
    if text in ('пропуск', 'missed', 'не был'):
        return Visit.Status.MISSED
    if text in ('отработка', 'makeup'):
        return Visit.Status.MAKEUP
    if text in ('заморозка', 'frozen'):
        return Visit.Status.FROZEN
    if text in ('пробный', 'trial'):
        return Visit.Status.TRIAL
    if text in ('отменен', 'отменён', 'cancelled', 'canceled'):
        return Visit.Status.CANCELLED
    return Visit.Status.PLANNED


def _normalize_status(value):
    return _text(value).lower().replace('ё', 'е').strip()


def _parse_date(value):
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_datetime(value):
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    else:
        text = _text(value)
        parsed = None
        for fmt in (
            '%d.%m.%Y %H:%M',
            '%d.%m.%y %H:%M',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%dT%H:%M',
            '%d.%m.%Y',
            '%d.%m.%y',
            '%Y-%m-%d',
        ):
            try:
                parsed = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if not parsed:
            return None
    if timezone.is_naive(parsed):
        return timezone.make_aware(parsed)
    return parsed


def _datetime_from_date(value):
    parsed_date = value or timezone.localdate()
    return timezone.make_aware(datetime.combine(parsed_date, time.min))


def _decimal(value, default=0):
    if _is_blank(value):
        return Decimal(default)
    try:
        return Decimal(str(value).replace(',', '.').replace(' ', ''))
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _int(value, default=0):
    if _is_blank(value):
        return default
    try:
        return int(_decimal(value, default=default))
    except (TypeError, ValueError):
        return default


def _text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_blank(value):
    return value is None or _text(value) == ''


def _join_notes(*values):
    return '\n'.join(_text(value) for value in values if not _is_blank(value))


def _visit_notes(row):
    ignored = ('фио', 'клиент', 'ученик', 'дата', 'статус', 'остаток')
    notes = []
    for key, value in row.items():
        if _is_blank(value) or any(token in key for token in ignored):
            continue
        notes.append(f'{key}: {_text(value)}')
    return '\n'.join(notes)


def _skip(result, sheet, row_number, reason):
    result['skipped'] += 1
    _warn(result, sheet, row_number, reason)


def _warn(result, sheet, row_number, message):
    result['warnings'].append(f'{sheet}, строка {row_number}: {message}')
